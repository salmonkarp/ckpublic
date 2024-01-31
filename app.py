from flask import *
from flask_mail import Mail, Message
from openpyxl import load_workbook
from PIL import Image
import os

#helper functions ========================================

#get image url from static folder
def get_image_url(product_name, images_folder):
    images_url_array = []
    for filename in os.listdir(images_folder):
        if filename.startswith(product_name + '@') and 'processed' not in filename:
            images_url_array.append('static/images/' + filename) 
    return images_url_array

#process all data from workbook before app is running
def process_excel_file():
    excel_file_path = 'static/products.xlsx'
    images_folder_path = 'static/images'
    wb = load_workbook(excel_file_path)
    
    sheet = wb['Products']
    products_list = []
    counter = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_name, full_price, half_price, description, *extra_elements = row
        image_url = get_image_url(product_name, images_folder_path)
        product_dict = {
            'id':counter,
            'name': product_name,
            'full_price': full_price,
            'half_price': half_price,
            'description': description,
            'image_url': image_url
        }
        products_list.append(product_dict)
        counter += 1
    
    sheet = wb['Hampers']
    hampers_list = []
    counter = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_name, price, description = row
        image_url = get_image_url(product_name, images_folder_path)
        product_dict = {
            'id':counter,
            'name': product_name,
            'price': price,
            'description': description,
            'image_url': image_url
        }
        hampers_list.append(product_dict)
        counter += 1
        
    sheet = wb['Snacks']
    snacks_list = []
    counter = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_name, price, description = row
        image_url = get_image_url(product_name, images_folder_path)
        product_dict = {
            'id':counter,
            'name': product_name,
            'price': price,
            'description': description,
            'image_url': image_url
        }
        snacks_list.append(product_dict)
        counter += 1
        
    sheet = wb['Signature']
    signature_list = []
    counter = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_name, price, description = row
        image_url = get_image_url(product_name, images_folder_path)
        product_dict = {
            'id':counter,
            'name': product_name,
            'price': price,
            'description': description,
            'image_url': image_url
        }
        signature_list.append(product_dict)
        counter += 1

    return products_list, hampers_list, snacks_list, signature_list

#get image url from cover folder
def get_all_cover_image_urls():
    cover_folder_path = 'static/cover'
    image_urls = []

    for filename in os.listdir(cover_folder_path):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')) and 'processed' not in filename:
            image_urls.append(cover_folder_path + '/' + filename)

    return image_urls

#process too high quality of an image
def process_image(image_path):
    updated_path = image_path.split('.')[0]
    ftype = image_path.split('.')[1]
    processed_image_path = f"{updated_path}_processed.{ftype}"
    try:
        img = Image.open(processed_image_path)
        return processed_image_path
    except:
        img = Image.open(image_path)
        desired_quality = 50
        img.save(processed_image_path, format='JPEG', quality=desired_quality)
        return processed_image_path

#jinja format currency
def format_currency(value):
    formatted_currency = "Rp {:,.0f}".format(value).replace(",", ".")
    return formatted_currency

#end of helper functions =================================



#global environments =====================================
app = Flask(__name__)
cover_image_urls = get_all_cover_image_urls()
products_list, hampers_list, snacks_list, signatures_list = process_excel_file()
app.jinja_env.filters['format_currency'] = format_currency
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'garyaxelmuliyono@gmail.com'
app.config['MAIL_PASSWORD'] = "LOL"
app.config['MAIL_DEFAULT_SENDER'] = 'garyaxelmuliyono@gmail.com'

mail = Mail(app)

#end of global environments===============================


#image processing routes==================================
@app.route('/static/images/<filename>')
def serve_image(filename):
    image_path = os.path.join('static', 'images', filename)
    max_file_size = 1 * 1024 * 1024
    if os.path.getsize(image_path) > max_file_size:
        processed_image_path = process_image(image_path)
    else:
        processed_image_path = image_path
    return send_file(processed_image_path)

@app.route('/static/cover/<filename>')
def serve_cover_image(filename):
    image_path = os.path.join('static', 'cover', filename)
    max_file_size = 2 * 1024 * 1024
    if os.path.getsize(image_path) > max_file_size:
        processed_image_path = process_image(image_path)
    else:
        processed_image_path = image_path
    return send_file(processed_image_path)

#end of image processing routes===========================


@app.route('/',methods=['GET'])
def root():
    return render_template('home.html', cover_image_urls = cover_image_urls)

@app.route('/products',methods=['GET'])
def signature():
    option = request.args.get('option','cookies')
    cover_image_urls = get_all_cover_image_urls()
    return render_template('products.html',products_list = products_list, hampers_list = hampers_list, snacks_list = snacks_list, signatures_list = signatures_list, option=option, cover_image_urls = cover_image_urls)

@app.route('/wheretobuy',methods=['GET'])
def wheretobuy():
    return render_template('wheretobuy.html', cover_image_urls = cover_image_urls)

@app.route('/contactus',methods=['GET','POST'])
def contactus():
    if request.method == 'GET':
        return render_template('contactus.html', cover_image_urls = cover_image_urls)
    else:
        full_name = request.form.get('fullName')
        email = request.form.get('email')
        message = request.form['message']
        subject = 'Contact Form Submission from Cookies Kingdom Website'
        body = f'Name: {full_name}\nEmail: {email}\nMessage: {message}'
        msg = Message(subject, recipients=['garyaxelmuliyono@gmail.com','monikamuliyono1@gmail.com'])
        msg.body = body
        try:
            mail.send(msg)
            response = {
                'script': 'document.getElementById("contact-form").reset(); alert("Message submitted successfully!");'
            }
            return jsonify(response)
        except Exception as e:
            print(e)
            response = {
                'script': f'alert("Error sending email: {str(e)}");'
            }
            return jsonify(response)

@app.route('/certifications',methods=['GET'])
def certifications():
    return render_template('certifications.html',cover_image_urls = cover_image_urls)

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)